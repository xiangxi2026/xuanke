"""
从Excel提取学生数据，生成系统所需的数据文件
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import json
import openpyxl

wb = openpyxl.load_workbook('../高一物理化学成绩等第_含身份证.xlsx')
ws = wb.active

students = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    exam_id = row[0]
    name = row[1]
    class_id = row[2]
    phy_score = row[3]
    phy_grade = row[4]
    chem_score = row[5]
    chem_grade = row[6]
    id_card = row[7]
    
    if not name or not id_card:
        continue
    
    # 清理身份证，去掉可能的"重名"前缀
    id_str = str(id_card).strip()
    if id_str.startswith('重名'):
        continue  # 跳过未处理的行
    
    # 判断是否可以报物化技
    phy_g = str(phy_grade).strip().upper() if phy_grade else ''
    chem_g = str(chem_grade).strip().upper() if chem_grade else ''
    can_wuhua = phy_g in 'ABCD' and chem_g in 'ABCD'
    
    students.append({
        'examId': str(exam_id).strip() if exam_id else '',
        'name': name.strip(),
        'class': str(class_id).strip() if class_id else '',
        'phyGrade': phy_g,
        'chemGrade': chem_g,
        'idCard': id_str,
        'canWuhua': can_wuhua
    })

# 去重（同姓名同身份证只保留一次）
seen = set()
unique_students = []
for s in students:
    key = (s['name'], s['idCard'])
    if key not in seen:
        seen.add(key)
        unique_students.append(s)

print(f'共提取 {len(unique_students)} 名学生')
print(f'可报"物化技": {sum(1 for s in unique_students if s["canWuhua"])}人')
print(f'不可报"物化技": {sum(1 for s in unique_students if not s["canWuhua"])}人')

# 保存学生数据（密码模式：仅身份证后6位作为验证）
with open('data/students.json', 'w', encoding='utf-8') as f:
    json.dump(unique_students, f, ensure_ascii=False, indent=2)

# 初始化报名数据
initial_data = {
    'counts': {
        '政史地': 0,
        '生地技': 0,
        '物化技': 0
    },
    'limits': {
        '政史地': 86,
        '生地技': 84,
        '物化技': 36
    },
    'submissions': {}
}

with open('data/submissions.json', 'w', encoding='utf-8') as f:
    json.dump(initial_data, f, ensure_ascii=False, indent=2)

print('数据已保存到 data/students.json 和 data/submissions.json')